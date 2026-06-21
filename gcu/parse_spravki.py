# -*- coding: utf-8 -*-
"""
Parse all supplementary справки files for a given ГЦУ date and load into
the spravki_* tables.

Usage (inside gcu-watch or gcu-mcp):
    python parse_spravki.py --date 2026-03-12 --dir /data/spravki/2026-03-12/

Files expected in --dir:
    Справка о наличии задержанных поездов.xlsx
    Суточная оперативная справка о случаях отказов в работе технических средств.xlsx
    Справка Локомотивы.xlsx
    Справка о работе припортовых станций на ДВОСТ ж.д. *.xlsx
    Справка о работе припортовых станций на ОКТ ж.д. *.xlsx
    Справка о работе припортовых станций на СКАВ ж.д. *.xlsx
    Справка Анализ выполнения участковой скорости*.xlsb
    Справка о Выполнении технической скорости*.xlsb
"""
import os, re, sys, argparse, datetime
import openpyxl

try:
    import pyxlsb
    HAS_XLSB = True
except ImportError:
    HAS_XLSB = False

DB_URL = os.environ.get("GCU_DATABASE_URL",
                        "postgresql://postgres:Gcu2026!@gcu-postgres:5432/postgres")

ROADS_ROW = ["СЕТЬ","ОКТ","КЛГ","МСК","ГОР","СЕВ","СКВ","ЮВС","ПРВ","КБШ","СВР","ЮУР","ЗСБ","КРС","ВСБ","ЗАБ","ДВС"]


def _clean(v):
    if v is None: return None
    s = str(v).strip().replace('\xa0', ' ')
    return s if s else None


def _num(v):
    if v is None: return None
    try: return float(v)
    except: return None


# ---------------------------------------------------------------------------
# 1. Задержанные поезда
# ---------------------------------------------------------------------------
def parse_delays(path, report_date):
    """Parse 'Справка о наличии задержанных поездов.xlsx'.

    Layout: row5 = road headers, then pairs of rows (п-да / ваг) per code.
    Returns list of dicts for spravki_delays.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # find road header row (contains 'Сеть' or 'СЕТЬ')
    header_idx = None
    for i, row in enumerate(rows):
        if any(str(v).strip().upper() in ('СЕТЬ', 'СЕТ') for v in row if v):
            header_idx = i
            break
    if header_idx is None:
        return []

    # road codes from header row — all non-empty cells after col0
    header_row = rows[header_idx]
    road_cols = {}  # col_index -> road_code
    for ci, v in enumerate(header_row):
        s = _clean(v)
        if s and ci > 0:
            road_cols[ci] = s.upper().strip()

    records = []
    code = None
    code_name = None
    train_row = None

    for row in rows[header_idx + 1:]:
        first = _clean(row[0]) if len(row) > 0 else None
        second = _clean(row[1]) if len(row) > 1 else None

        if not first and not second:
            continue

        # detect code row: starts with ˈˈXˈˈ or «XX» - Название
        m = re.match(r"[«\'\ˈ\"ˈ]+([\d]+)[»\'\ˈ\"ˈ]+\s*[-–—]\s*(.+)", first or "")
        if m:
            code = m.group(1).strip()
            code_name = m.group(2).strip()
            # check if п-да is on same row or next
            if second == 'п-да' or second == 'п-дa':
                train_row = row
            else:
                train_row = None
            continue

        if first == 'ВСЕГО':
            code = 'ВСЕГО'
            code_name = 'Итого все коды'
            if second == 'п-да' or second == 'п-дa':
                train_row = row
            continue

        # п-да row
        if second in ('п-да', 'п-дa') or (first is None and second and 'п' in str(second).lower()):
            train_row = row
            continue

        # ваг row — follows п-да
        if second == 'ваг' or (first is None and train_row is not None):
            wagon_row = row
            # emit records for each road
            for ci, road in road_cols.items():
                trains = None
                wagons = None
                if train_row is not None and ci < len(train_row):
                    trains = int(train_row[ci]) if isinstance(train_row[ci], (int, float)) else None
                if ci < len(wagon_row):
                    wagons = int(wagon_row[ci]) if isinstance(wagon_row[ci], (int, float)) else None
                if trains is not None or wagons is not None:
                    records.append(dict(
                        report_date=report_date, delay_code=code,
                        delay_name=code_name, road_code=road,
                        trains=trains, wagons=wagons
                    ))
            train_row = None
            continue

    return records


# ---------------------------------------------------------------------------
# 2. Отказы технических средств
# ---------------------------------------------------------------------------
def parse_failures(path, report_date):
    """Parse 'Суточная оперативная справка о случаях отказов.xlsx'.

    Layout: row8+ = dept, 2025, 2026, +/-%, resolved, registered, investigated
    Returns list of dicts for spravki_failures.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # data starts after the row containing column numbers 1,2,3,4,5,6,7
    data_start = 0
    for i, row in enumerate(rows):
        nums = [v for v in row if isinstance(v, (int, float)) and 1 <= v <= 7]
        if len(nums) >= 4:
            data_start = i + 1
            break

    records = []
    section = None   # 1=произошедшие на территории дороги, 2=по ответственности, 3=по комплексам
    for row in rows[data_start:]:
        # Label sits in col[0] for road rows, but col[1] for complex rows
        # (локомотивный/инфраструктурный комплекс) where col[0] is empty.
        dept = _clean(row[0]) if len(row) > 0 and row[0] else None
        if not dept and len(row) > 1 and row[1]:
            dept = _clean(row[1])
        if not dept:
            continue
        # «РазделN…» header rows mark which cut of the data follows. The source
        # repeats the SAME roads under Раздел1 (где произошёл отказ) and Раздел2
        # (по чьей ответственности) — each block already sums to the network ИТОГО,
        # so they must NEVER be summed together. Stamp every following row with the
        # section so callers can filter to one block. Раздел3 = разбивка по комплексам.
        m = re.match(r'Раздел\s*([123])', dept)
        if m:
            section = int(m.group(1))
            continue
        if dept.startswith('ИТОГО') and len(dept) > 20:
            continue
        # skip section headers (all caps long strings)
        if dept.isupper() and len(dept) > 30:
            continue

        def gi(i):
            v = row[i] if i < len(row) else None
            if isinstance(v, (int, float)): return int(v) if v == int(v) else v
            return None

        def gf(i):
            v = row[i] if i < len(row) else None
            # Source mixes formats: text "38,46%" vs float fraction 0.322.
            # Excel stores %-formatted cells as fractions → scale to percent.
            if isinstance(v, float):
                return round(v * 100, 2)
            try: return float(str(v).replace('%','').replace(',','.').strip())
            except: return None

        def gnum(i):
            """Raw float (hours/poezdo-hours) — NOT scaled like gf()'s percent."""
            v = row[i] if i < len(row) else None
            if isinstance(v, (int, float)):
                return round(float(v), 2)
            try: return round(float(str(v).replace(',', '.').strip()), 2)
            except: return None

        # The dept name cell is merged across col[0:1], so the numeric block
        # always starts at index [2]. Full source layout (cols, 0-based):
        #   2=отказы 2025, 3=отказы 2026, 4=+/-%, 5=устранено, 6=принято к учёту,
        #   7=расследовано, 8=на расследовании,
        #   9=продолж.отказов 2025 (час), 10=продолж.отказов 2026 (час), 11=+/-%,
        #   16=задержано грузовых поездов (кол-во), 17=продолж.задержки грузовых (поездо-час)
        f25, f26 = gi(2), gi(3)
        pct = gf(4)
        resolved, registered, investigated = gi(5), gi(6), gi(7)
        dur25, dur26 = gnum(9), gnum(10)
        dur_pct = gf(11)
        freight_trains_delayed = gi(16)
        freight_train_hours = gnum(17)

        if f25 is None and f26 is None:
            continue

        records.append(dict(
            report_date=report_date, dept=dept, section=section,
            failures_2025=f25, failures_2026=f26,
            change_pct=pct, resolved=resolved,
            registered=registered, investigated=investigated,
            duration_2025=dur25, duration_2026=dur26, duration_change_pct=dur_pct,
            freight_trains_delayed=freight_trains_delayed,
            freight_train_hours=freight_train_hours,
        ))
    return records


# ---------------------------------------------------------------------------
# 3. Локомотивы
# ---------------------------------------------------------------------------
def parse_locomotives(path, report_date):
    """Parse 'Справка Локомотивы.xlsx'.

    Each data row carries ALL THREE traction types side-by-side (not split by
    section headers): Переменный ток (cols 1-10), Постоянный ток (11-20),
    Тепловозная тяга (21-30). Within each block:
      [+0] План всего · [+2] План груз · [+3] Факт всего · [+5] Факт груз ·
      [+6] +/- всего · [+8] +/- груз · [+9] Резерв
    The old parser read only the AC-block груз columns and labelled everything
    section='AC' — dropping DC + diesel entirely and the Резерв/всего-парк. Now we
    emit ONE record per (road, traction) with всего + груз + delta + reserve.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    BLOCKS = [('AC', 1), ('DC', 11), ('diesel', 21)]  # (traction, base col)
    records = []
    current_polygon = None
    POLYGON_KEYWORDS = ['юго-зап', 'северо-зап', 'восточн', 'западн', 'московск', 'октябрьск']

    def g(i):
        v = row[i] if i < len(row) else None
        return int(v) if isinstance(v, (int, float)) else None

    for row in rows:
        first = _clean(row[0]) if len(row) > 0 and row[0] else None
        if not first:
            continue
        fl = first.lower()
        if fl.strip() in ('план', 'факт', '+/-', 'дорога'):
            continue
        if any(kw in fl for kw in ['содержание', 'анализ', 'таблиц', 'страниц',
                                   'справк', 'полигон', 'переменный', 'постоянный', 'тепловоз']):
            continue
        # polygon vs road: a polygon name is also a data row (has aggregate numbers)
        is_polygon = any(kw in fl for kw in POLYGON_KEYWORDS)
        if is_polygon:
            current_polygon = first.strip()

        road_name = first.strip()
        for traction, b in BLOCKS:
            plan_total  = g(b + 0)
            plan_freight= g(b + 2)
            fact_total  = g(b + 3)
            fact_freight= g(b + 5)
            delta_total = g(b + 6)
            delta_freight = g(b + 8)
            reserve     = g(b + 9)
            # skip an empty traction block for this road
            if plan_total is None and fact_total is None and plan_freight is None and fact_freight is None:
                continue
            records.append(dict(
                report_date=report_date, section=traction,
                polygon=current_polygon if not is_polygon else None,
                road=road_name,
                plan=plan_freight, fact=fact_freight, delta=delta_freight,  # back-compat (груз)
                plan_total=plan_total, fact_total=fact_total, delta_total=delta_total,
                reserve=reserve,
            ))
    return records


# ---------------------------------------------------------------------------
# 4. Припортовые станции
# ---------------------------------------------------------------------------
def parse_port_stations(path, report_date, road_code):
    """Parse 'Справка о работе припортовых станций на X ж.д..xlsx'.

    CRITICAL: rows are a 5-level HIERARCHY encoded as cell INDENTATION in col A
    (alignment.indent), which the data-only read discards. Without row_level the
    agent cannot tell a road-subtotal (ДАЛЬНЕВОСТОЧНАЯ) from a real station
    (Находка-Восточная) from a cargo line (Каменный уголь) — it double-counts and
    lists road totals as stations. We read formatting + values and map indent:
      1=network('ИТОГО ПО ПОРТАМ СЕТИ') 2=road 3=port(узел) 4=terminal 5=cargo.
    """
    # Two loads: data_only for VALUES, default for INDENT (openpyxl can't do both).
    wb_v = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb_v.active.iter_rows(values_only=True)); wb_v.close()
    wb_f = openpyxl.load_workbook(path)  # formatting (indent) — not read_only
    ws_f = wb_f.active
    indents = {}
    for i in range(1, ws_f.max_row + 1):
        c = ws_f.cell(i, 1)
        al = c.alignment
        indents[i - 1] = int(al.indent) if al and al.indent else 0  # 0-based row key
    wb_f.close()

    # Absolute indent is NOT consistent across the 3 road files: the ДВОСТ file
    # has an extra top "ИТОГО ПО ПОРТАМ СЕТИ" (network) row that shifts everything
    # down by 1, while ОКТ/СКАВ start at the road itself. So map level RELATIVE to
    # the file's minimum data indent, and pin network/road by NAME, not indent.
    REL = {1: "port", 2: "terminal", 3: "cargo", 4: "cargo"}  # indent above the road row
    # Full road names (UPPER) to recognise a road-subtotal row vs a station.
    ROAD_NAMES_UP = {
        "ОКТЯБРЬСКАЯ", "ДАЛЬНЕВОСТОЧНАЯ", "СЕВЕРО-КАВКАЗСКАЯ", "КАЛИНИНГРАДСКАЯ",
        "МОСКОВСКАЯ", "ГОРЬКОВСКАЯ", "СЕВЕРНАЯ", "ЮГО-ВОСТОЧНАЯ", "ПРИВОЛЖСКАЯ",
        "КУЙБЫШЕВСКАЯ", "СВЕРДЛОВСКАЯ", "ЮЖНО-УРАЛЬСКАЯ", "ЗАПАДНО-СИБИРСКАЯ",
        "КРАСНОЯРСКАЯ", "ВОСТОЧНО-СИБИРСКАЯ", "ЗАБАЙКАЛЬСКАЯ",
    }
    # Find the indent of the ROAD row — ports/terminals/cargo are measured from it.
    road_indent = None
    for ri in range(4, len(rows)):
        nm = rows[ri][0]
        if nm and str(nm).strip().upper() in ROAD_NAMES_UP:
            road_indent = indents.get(ri, 0); break
    if road_indent is None:  # no explicit road row → ports sit at file's min indent - 1
        di = [indents.get(ri, 0) for ri in range(4, len(rows)) if rows[ri][0]]
        road_indent = (min(di) - 1) if di else 0

    # FIXED column layout (0-based). IMPORTANT SEMANTICS (per source header, ДВС/ОКТ/СКАВ):
    #   [1]погрузка всего  [2]погрузка ср/сут
    #   "Наличие вагонов на 18:00" = вагоны НАЗНАЧЕНИЕМ на припортовые станции (НЕ на самой
    #   станции!): [3]на СЕТИ норма [4]на СЕТИ факт [5]отставл.поездов(сеть, назнач.→порты)
    #   [6]на ДОРОГЕ норма [7]на ДОРОГЕ факт [8]отставл.поездов(дорога назнач.)
    #   [9]НАЛИЧ. НА САМОЙ СТАНЦИИ на 18:00 12.03  ← это и есть «вагоны на станции»
    #   [10]перер.способность [11]ВЫГРУЗКА на 18:00 [12]выгрузка ср/сут [13]выгрузка на 06:00
    #   [14]НАЛИЧ. НА САМОЙ СТАНЦИИ на 06:00 13.03  [15]план выгрузки на 18:00 13.03
    records = []
    for ri, row in enumerate(rows[4:], start=4):
        station = _clean(row[0]) if len(row) > 0 and row[0] else None
        if not station or len(station) < 3:
            continue
        if any(w in station.lower() for w in ['справ', 'погруз', 'налич', 'норма', 'перер']):
            continue

        def g(i):
            return _num(row[i]) if i < len(row) and isinstance(row[i], (int, float)) else None

        # Level: network/road pinned by name; deeper levels by indent above road row.
        st_up = station.upper()
        rel = indents.get(ri, road_indent) - road_indent
        if "ИТОГО ПО ПОРТАМ" in st_up:
            row_level = "network"
        elif st_up in ROAD_NAMES_UP:
            row_level = "road"
        else:
            row_level = REL.get(rel, "cargo" if rel >= 3 else "port")

        load_total   = g(1)    # ПОГРУЗКА всего за сутки (НЕ план выгрузки)
        load_avg     = g(2)    # погрузка ср/сут
        capacity     = g(10)   # перерабатывающая способность
        load_fact    = g(11)   # ВЫГРУЗКА факт на 18:00
        unload_avg   = g(12)   # выгрузка ср/сут
        # «Наличие вагонов на 18:00» = вагоны НАЗНАЧЕНИЕМ на припортовые станции,
        # а НЕ вагоны на самой станции. wagons_total/road имена сохранены (на них
        # завязан view v_ports_network) — смысл уточнён в комментариях колонок.
        wagons_total = g(4)    # наличие на СЕТИ назнач.→порты, факт
        wagons_road  = g(7)    # наличие на ДОРОГЕ назнач.→порты, факт
        wagons_dest_net_norm  = g(3)   # на СЕТИ назнач.→порты, НОРМА
        wagons_dest_road_norm = g(6)   # на ДОРОГЕ назнач.→порты, НОРМА
        wagons_at_station_18 = g(9)    # НАЛИЧ. НА САМОЙ СТАНЦИИ на 18:00 (это «вагоны на станции»)
        wagons_at_station_06 = g(14)   # НАЛИЧ. НА САМОЙ СТАНЦИИ на 06:00 след. суток
        detained     = g(5)    # отставленных поездов (сеть, назнач.→порты)
        detained_road= g(8)    # отставленных поездов (на дороге назнач.)
        unload_06    = g(13)   # выгрузка на 06:00 след. суток
        unload_plan_next = g(15)  # план выгрузки на 18:00 следующих суток

        if load_fact is None and capacity is None and wagons_total is None:
            continue

        records.append(dict(
            report_date=report_date, road=road_code, station=station,
            row_level=row_level,
            load_total=load_total, load_fact=load_fact, capacity=capacity,
            load_avg=load_avg, unload_avg=unload_avg,
            wagons_total=int(wagons_total) if wagons_total else None,
            wagons_road=int(wagons_road) if wagons_road else None,
            wagons_dest_net_norm=int(wagons_dest_net_norm) if wagons_dest_net_norm else None,
            wagons_dest_road_norm=int(wagons_dest_road_norm) if wagons_dest_road_norm else None,
            wagons_at_station_18=int(wagons_at_station_18) if wagons_at_station_18 else None,
            wagons_at_station_06=int(wagons_at_station_06) if wagons_at_station_06 else None,
            detained_trains=int(detained) if detained else None,
            detained_trains_road=int(detained_road) if detained_road else None,
            unload_06=unload_06,
            unload_plan_next=unload_plan_next,
        ))
    return records


# ---------------------------------------------------------------------------
# 5. Скорость (участковая + техническая)
# ---------------------------------------------------------------------------
def parse_speed_xlsb(path, report_date, speed_type):
    """Parse xlsb speed справки. Returns list of dicts for spravki_speed."""
    if not HAS_XLSB:
        print("  pyxlsb not available — skipping", path)
        return []

    records = []
    with pyxlsb.open_workbook(path) as wb:
        for sh in wb.sheets:
            with wb.get_sheet(sh) as ws:
                all_rows = list(ws.rows())
                # find data rows: first cell is a road name (string, not header)
                for row in all_rows:
                    vals = [c.v for c in row]
                    first = _clean(vals[0]) if vals else None
                    if not first or len(first) < 3:
                        continue
                    # skip title/header rows
                    if any(kw in first.lower() for kw in ['анализ','выполнен','справ','дорога','таблиц','страниц','гвц','за от']):
                        continue
                    numerics = [_num(c.v) for c in row[1:] if isinstance(c.v, (int, float))]
                    if len(numerics) < 3:
                        continue

                    road = first.strip()
                    # FIXED columns by index (cell.c). Main block 1-7 is gruzovoe
                    # (what answers are based on). Section files ALSO carry two
                    # sub-variants the old collapse-parser dropped:
                    #   8-11  = «без передаточных и вывозных поездов»
                    #   12-15 = «передаточные и вывозные поезда»
                    # Technical files have only the 7-col block → extras stay None.
                    cv = {c.c: _num(c.v) for c in row if isinstance(c.v, (int, float))}
                    def col(i): return cv.get(i)
                    records.append(dict(
                        report_date=report_date, speed_type=speed_type, road=road,
                        prev_year=col(1), norm=col(2), day_fact=col(3), day_delta=col(4),
                        month_fact=col(5), month_delta=col(6), year_delta=col(7),
                        # sub-variant: без передаточных/вывозных
                        nopass_prev_year=col(8), nopass_day_fact=col(9),
                        nopass_month_fact=col(10), nopass_year_delta=col(11),
                        # sub-variant: передаточные/вывозные
                        pass_prev_year=col(12), pass_day_fact=col(13),
                        pass_month_fact=col(14), pass_year_delta=col(15),
                    ))
    return records


# ---------------------------------------------------------------------------
# DB write helpers
# ---------------------------------------------------------------------------
def _delete_date(conn, table, report_date):
    with conn.cursor() as cur:
        cur.execute(f"DELETE FROM {table} WHERE report_date = %s", (report_date,))


def write_delays(conn, records):
    sql = """INSERT INTO spravki_delays
             (report_date,delay_code,delay_name,road_code,trains,wagons)
             VALUES (%s,%s,%s,%s,%s,%s)"""
    with conn.cursor() as cur:
        for r in records:
            cur.execute(sql, (r['report_date'],r['delay_code'],r['delay_name'],
                              r['road_code'],r['trains'],r['wagons']))


def write_failures(conn, records):
    sql = """INSERT INTO spravki_failures
             (report_date,dept,section,failures_2025,failures_2026,change_pct,resolved,registered,investigated,
              duration_2025,duration_2026,duration_change_pct,freight_trains_delayed,freight_train_hours)
             VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
    with conn.cursor() as cur:
        for r in records:
            cur.execute(sql, (r['report_date'],r['dept'],r.get('section'),r['failures_2025'],r['failures_2026'],
                              r['change_pct'],r['resolved'],r['registered'],r['investigated'],
                              r.get('duration_2025'),r.get('duration_2026'),r.get('duration_change_pct'),
                              r.get('freight_trains_delayed'),r.get('freight_train_hours')))


def write_locomotives(conn, records):
    sql = """INSERT INTO spravki_locomotives
             (report_date,section,polygon,road,plan,fact,delta,
              plan_total,fact_total,delta_total,reserve)
             VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
    with conn.cursor() as cur:
        for r in records:
            cur.execute(sql, (r['report_date'],r['section'],r['polygon'],r['road'],
                              r['plan'],r['fact'],r['delta'],
                              r.get('plan_total'),r.get('fact_total'),
                              r.get('delta_total'),r.get('reserve')))


def write_port_stations(conn, records):
    sql = """INSERT INTO spravki_port_stations
             (report_date,road,station,row_level,load_total,load_fact,capacity,
              load_avg,unload_avg,wagons_total,wagons_road,detained_trains,
              detained_trains_road,unload_plan_next,
              wagons_dest_net_norm,wagons_dest_road_norm,
              wagons_at_station_18,wagons_at_station_06,unload_06)
             VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
    with conn.cursor() as cur:
        for r in records:
            cur.execute(sql, (r['report_date'],r['road'],r['station'],r.get('row_level'),
                              r['load_total'],r['load_fact'],r['capacity'],
                              r.get('load_avg'),r.get('unload_avg'),r['wagons_total'],
                              r['wagons_road'],r['detained_trains'],
                              r.get('detained_trains_road'),r.get('unload_plan_next'),
                              r.get('wagons_dest_net_norm'),r.get('wagons_dest_road_norm'),
                              r.get('wagons_at_station_18'),r.get('wagons_at_station_06'),
                              r.get('unload_06')))


def write_speed(conn, records):
    sql = """INSERT INTO spravki_speed
             (report_date,speed_type,road,prev_year,norm,day_fact,day_delta,month_fact,month_delta,year_delta,
              nopass_prev_year,nopass_day_fact,nopass_month_fact,nopass_year_delta,
              pass_prev_year,pass_day_fact,pass_month_fact,pass_year_delta)
             VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
    with conn.cursor() as cur:
        for r in records:
            cur.execute(sql, (r['report_date'],r['speed_type'],r['road'],
                              r['prev_year'],r['norm'],r['day_fact'],r['day_delta'],
                              r['month_fact'],r['month_delta'],r['year_delta'],
                              r.get('nopass_prev_year'),r.get('nopass_day_fact'),
                              r.get('nopass_month_fact'),r.get('nopass_year_delta'),
                              r.get('pass_prev_year'),r.get('pass_day_fact'),
                              r.get('pass_month_fact'),r.get('pass_year_delta')))


# ---------------------------------------------------------------------------
# Main ingest
# ---------------------------------------------------------------------------
def ingest_dir(dirpath, report_date_str, force=False):
    import psycopg
    report_date = datetime.date.fromisoformat(report_date_str)
    conn = psycopg.connect(DB_URL)

    try:
        # ── задержанные ──
        f = _find(dirpath, 'задержанных поездов')
        if f:
            print(f"  [delays] {os.path.basename(f)}")
            _delete_date(conn, 'spravki_delays', report_date)
            recs = parse_delays(f, report_date)
            write_delays(conn, recs)
            print(f"    → {len(recs)} rows")
        else:
            print("  [delays] not found — skip")

        # ── отказы ──
        f = _find(dirpath, 'отказов в работе')
        if f:
            print(f"  [failures] {os.path.basename(f)}")
            _delete_date(conn, 'spravki_failures', report_date)
            recs = parse_failures(f, report_date)
            write_failures(conn, recs)
            print(f"    → {len(recs)} rows")
        else:
            print("  [failures] not found — skip")

        # ── локомотивы ──
        f = _find(dirpath, 'локомотив', ci=True)
        if f:
            print(f"  [locomotives] {os.path.basename(f)}")
            _delete_date(conn, 'spravki_locomotives', report_date)
            recs = parse_locomotives(f, report_date)
            write_locomotives(conn, recs)
            print(f"    → {len(recs)} rows")
        else:
            print("  [locomotives] not found — skip")

        # ── припортовые ──
        # The port справки come one-file-per-road, but vendors name the road in
        # the filename inconsistently: short code (ДВОСТ/ОКТ/СКАВ) in some
        # exports, full road name (Дальневосточная/Октябрьская/Северо-Кавказская)
        # in others. Match BOTH spellings so a naming change doesn't silently
        # skip a road. The regex alternation is anchored on 'припортовых'.
        _delete_date(conn, 'spravki_port_stations', report_date)
        port_roads = [
            ('ДВОСТ', ['двост', 'дальневосточн']),
            ('ОКТ',   ['окт', 'октябрьск']),
            ('СКАВ',  ['скав', 'северо-кавказск', 'сев-кавказск']),
        ]
        for road_code, kws in port_roads:
            alt = "|".join(kws)
            f = _find(dirpath, f'припортовых.*(?:{alt})')
            if f:
                print(f"  [ports:{road_code}] {os.path.basename(f)}")
                recs = parse_port_stations(f, report_date, road_code)
                write_port_stations(conn, recs)
                print(f"    → {len(recs)} rows")
            else:
                print(f"  [ports:{road_code}] not found — skip")

        # ── скорость xlsb ──
        _delete_date(conn, 'spravki_speed', report_date)
        for kw, stype in [('участков', 'section'), ('технической скорости', 'technical')]:
            f = _find(dirpath, kw, exts=['.xlsb'])
            if f:
                print(f"  [speed:{stype}] {os.path.basename(f)}")
                recs = parse_speed_xlsb(f, report_date, stype)
                write_speed(conn, recs)
                print(f"    → {len(recs)} rows")
            else:
                print(f"  [speed:{stype}] not found — skip")

        # ── сортировочные станции xlsb (отдельный парсер parse_sort_stations) ──
        # Раньше это была РУЧНАЯ операция (parse_sort_stations.py запускали отдельно),
        # из-за чего для новых дат таблица молча оставалась пустой. Теперь встроено.
        f = _find(dirpath, 'сортировочн', exts=['.xlsb'])
        if f:
            try:
                import parse_sort_stations as PSS
                recs = PSS.parse(f, report_date_str)
                _delete_date(conn, 'spravki_sort_stations', report_date)
                with conn.cursor() as cur:
                    for r in recs:
                        cur.execute(
                            "INSERT INTO spravki_sort_stations "
                            "(report_date, road, station, period, working_park, rosp_total, "
                            " arrived_trains, refused_trains, rosp_no_pere, formed_trains, "
                            " sent_trains, avg_weight, avg_length, park_norm, "
                            " idle_transit_pere, idle_transit_pere_norm, "
                            " idle_transit_nopere, idle_transit_nopere_norm, raw_extra) "
                            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb)",
                            (r["report_date"], r["road"], r["station"], r["period"],
                             r.get("working_park"), r.get("rosp_total"),
                             r.get("arrived_trains"), r.get("refused_trains"),
                             r.get("rosp_no_pere"), r.get("formed_trains"),
                             r.get("sent_trains"), r.get("avg_weight"),
                             r.get("avg_length"), r.get("park_norm"),
                             r.get("idle_transit_pere"), r.get("idle_transit_pere_norm"),
                             r.get("idle_transit_nopere"), r.get("idle_transit_nopere_norm"),
                             r.get("raw_extra")))
                print(f"  [sort_stations] {os.path.basename(f)}\n    → {len(recs)} rows")
            except Exception as e:
                print(f"  [sort_stations] FAILED: {e} — skip (не ломает остальное)")
        else:
            print("  [sort_stations] not found — skip")

        # ── ограничения скорости ──
        # ВНИМАНИЕ: исходный xlsx хранит данные КАРТИНКОЙ (PNG), автопарсера нет.
        # Данные грузятся вручную через load_speed_restrictions.py из JSON. Здесь
        # только предупреждаем, чтобы для новой даты не забыли (молчаливый пропуск = баг).
        f = _find(dirpath, 'ограничени', exts=['.xlsx'])
        if f:
            have = False
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM spravki_speed_restrictions WHERE report_date=%s LIMIT 1",(report_date,))
                have = cur.fetchone() is not None
            if not have:
                print(f"  [speed_restrictions] ВНИМАНИЕ: файл есть, но данных за {report_date_str} "
                      f"в БД НЕТ. Источник — картинка (PNG) в xlsx; загрузи вручную через "
                      f"load_speed_restrictions.py (JSON). Иначе вопросы об ограничениях вернут «нет данных».")
            else:
                print(f"  [speed_restrictions] данные за {report_date_str} уже загружены (JSON-loader)")

        conn.commit()
        print(f"\n  [OK] ingest complete for {report_date_str}")
    except Exception as e:
        conn.rollback()
        import traceback; traceback.print_exc()
        raise
    finally:
        conn.close()


def _find(dirpath, pattern, ci=False, exts=None):
    """Find file matching regex pattern in dirpath."""
    flags = re.IGNORECASE if ci else 0
    pat = re.compile(pattern, re.IGNORECASE)
    for fname in os.listdir(dirpath):
        if exts and not any(fname.lower().endswith(e) for e in exts):
            continue
        if pat.search(fname):
            return os.path.join(dirpath, fname)
    return None


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--date', required=True, help='YYYY-MM-DD')
    ap.add_argument('--dir', required=True, help='directory with справки files')
    ap.add_argument('--force', action='store_true')
    args = ap.parse_args()
    print(f"Ingesting справки for {args.date} from {args.dir}")
    ingest_dir(args.dir, args.date, args.force)
