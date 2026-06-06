# -*- coding: utf-8 -*-
"""
ГЦУ daily report parser — v2. Writes to the typed-column schema (reports,
report_sheets, metrics, investment_metrics, report_comments), NOT the legacy
JSONB-everything prototype.

Idempotent by sha256: re-running on the same file is a no-op. Re-running on a
modified file (different hash) inserts a new report and leaves the old one.
To force-reload a date, use --force (deletes the existing report rows for that
date first; ON DELETE CASCADE cleans up dependents).

Usage:
    python parse_gtsu_v2.py <file.xlsx> [--date YYYY-MM-DD] [--force]
    python parse_gtsu_v2.py <directory>  [--force]      # batch mode

Sheet types:
  • "Доклад Ц ЦЗ" + "Срок доставки"  → metrics (operational, typed)
  • "Инвест" + "Инвест Факт"          → investment_metrics
"""
import os
import re
import sys
import argparse
import datetime as dt
import hashlib
import io

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl


# ---------------------------------------------------------------------------
# DB driver shim
# ---------------------------------------------------------------------------
def connect():
    """Connect honoring (in order): GCU_DATABASE_URL → libpq PG* env vars →
    a localhost default. The PG* fallback matches what gcu-watch/gcu-upload
    set in docker-compose, so the same parser works in either container."""
    import psycopg
    dsn = os.environ.get("GCU_DATABASE_URL")
    if dsn:
        return psycopg.connect(dsn, connect_timeout=10)
    if os.environ.get("PGHOST"):
        # libpq picks up PGHOST/PGPORT/PGUSER/PGPASSWORD/PGDATABASE automatically
        return psycopg.connect(connect_timeout=10)
    return psycopg.connect(
        "postgresql://postgres:Gcu2026!@127.0.0.1:5432/postgres",
        connect_timeout=10,
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def clean(v):
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    return s or None


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def item_depth(num_str):
    if not num_str:
        return None
    return len([p for p in str(num_str).split(".") if p.strip()])


def sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


# Filename date inference: ГЦУ-2026-04-01 or ГЦУ-03-09 (short → assume 2022)
_DATE_LONG = re.compile(r"(\d{4})[-_.](\d{2})[-_.](\d{2})")
_DATE_SHORT = re.compile(r"(\d{2})[-_.](\d{2})\.xlsx?$", re.I)


def infer_date(path, override=None):
    if override:
        return dt.date.fromisoformat(override)
    name = os.path.basename(path)
    m = _DATE_LONG.search(name)
    if m:
        return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = _DATE_SHORT.search(name)
    if m:
        return dt.date(2022, int(m.group(1)), int(m.group(2)))
    return dt.date.today()


# Baseline-year: read from header cell (row 5, col 10 contains "к 2021" or "к 2025")
def detect_baseline_year(wb):
    for sheet in ("Доклад Ц ЦЗ", "Срок доставки"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        # column J = 10
        for r in (5, 6):
            v = ws.cell(row=r, column=10).value
            if v and isinstance(v, str):
                m = re.search(r"к\s*(\d{4})", v)
                if m:
                    return int(m.group(1))
    return None


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------
SECTION_HEADER_RE = re.compile(r"^([IVX]+)[.\s\xa0]+(.+)$")
NUMBERED_HEADER_RE = re.compile(r"^(\d{1,2})[.\s]+(.+)$")
PRIORITY_FLAG_RE = re.compile(r"\*|приоритет", re.I)


# Roads we'll detect in the "indicator" column (leaf rows)
RUSSIAN_ROADS = {
    "Октябрьская", "Калининградская", "Московская", "Горьковская", "Северная",
    "Северо-Кавказская", "Юго-Восточная", "Приволжская", "Куйбышевская",
    "Свердловская", "Южно-Уральская", "Западно-Сибирская", "Красноярская",
    "Восточно-Сибирская", "Забайкальская", "Дальневосточная", "Сахалинская",
    "Северо-Западная", "Крымская",
}


def is_road(name):
    if not name:
        return None
    s = name.strip()
    # match e.g. "Октябрьская", "Октябрьская ж.д.", "Октябрьская жд"
    for road in RUSSIAN_ROADS:
        if s.startswith(road) or road in s.split():
            return road
    return None


def parse_operational(ws, sheet_name, section_code, header_row=4, data_start=6):
    """Parse Доклад Ц ЦЗ / Срок доставки → list of metric dicts + comment dicts.

    Column layout (0-based from col B=index 1):
      B=item_number   C=priority(*)  D=name  E=unit  F=zone  G=responsible
      H=day_fact     I=day_to_plan   J=day_to_prev_year
      K=month_fact   L=month_to_plan M=month_to_prev_yr
      N=year_fact    O=year_to_plan  P=year_to_prev_yr
      Q=commentary   R=management_action
    """
    metrics, comments = [], []
    current_category = None
    stack = []  # (depth, num, name) — for category inheritance only

    for r in range(data_start, ws.max_row + 1):
        get = lambda c: ws.cell(row=r, column=c).value
        num_str = clean(get(2))
        prio = clean(get(3))
        name = clean(get(4))
        unit = clean(get(5))
        zone = get(6)
        resp = clean(get(7))
        d_f, d_p, d_y = num(get(8)), num(get(9)), num(get(10))
        m_f, m_p, m_y = num(get(11)), num(get(12)), num(get(13))
        y_f, y_p, y_y = num(get(14)), num(get(15)), num(get(16))
        comm = clean(get(17))
        action = clean(get(18))

        if not name:
            continue

        # Section header (Roman): "II. МОНИТОРИНГ СРОКА…" — sets category, no metric
        if SECTION_HEADER_RE.match(name) and not num_str and all(
            v is None for v in (d_f, d_p, d_y, m_f, m_p, m_y, y_f, y_p, y_y)
        ):
            current_category = name
            stack = []
            continue

        # Top-level numbered header: "1. ГРУЗОВЫЕ ПЕРЕВОЗКИ" — category set, no metric
        if num_str and re.match(r"^\d+$", num_str) and all(
            v is None for v in (d_f, d_p, d_y, m_f, m_p, m_y, y_f, y_p, y_y)
        ):
            current_category = name
            stack = [(1, num_str, name)]
            continue

        depth = item_depth(num_str)
        if depth:
            stack = [s for s in stack if s[0] < depth]
            stack_parent_name = stack[-1][2] if stack else None
        else:
            stack_parent_name = None

        # Filter clearly non-metric rows (sections, totals without numbers)
        has_any_value = any(v is not None for v in (d_f, d_p, d_y, m_f, m_p, m_y, y_f, y_p, y_y))
        has_comment = bool(comm or action)
        if not has_any_value and not has_comment:
            if depth:
                stack.append((depth, num_str, name))
            continue

        zone_int = None
        if isinstance(zone, (int, float)):
            z = int(zone)
            if z in (0, 1, 2, 4):
                zone_int = z

        road = is_road(name)

        # parent_indicator = IMMEDIATE parent of this indicator_number, derived
        # from the dotted string itself (e.g. '7.7.3.3.1' → '7.7.3.3'). This is
        # unambiguous and survives sheet ordering quirks; the stack-based
        # version conflated rows whose immediate parent wasn't itself a data row.
        parent_num = None
        if num_str and "." in num_str:
            parent_num = num_str.rsplit(".", 1)[0]

        # populates: which period actually carries data for this indicator.
        has_d = any(v is not None for v in (d_f, d_p, d_y))
        has_m = any(v is not None for v in (m_f, m_p, m_y))
        has_y = any(v is not None for v in (y_f, y_p, y_y))
        if has_d and not has_m and not has_y:
            populates = "daily"
        elif has_m and not has_d and not has_y:
            populates = "monthly"
        elif has_y and not has_d and not has_m:
            populates = "yearly"
        elif has_d or has_m or has_y:
            populates = "mixed"
        else:
            populates = "none"

        metrics.append(dict(
            indicator_number=num_str,
            parent_indicator=parent_num,
            is_priority=bool(prio and "*" in prio),
            section_roman=section_code,
            name=name,
            category=current_category if current_category and (
                not SECTION_HEADER_RE.match(current_category)
            ) else (stack_parent_name if depth and depth >= 2 else current_category),
            road=road,
            unit=unit,
            responsible=resp,
            zone=zone_int,
            day_fact=d_f, day_to_plan=d_p, day_to_prev_year=d_y,
            month_fact=m_f, month_to_plan=m_p, month_to_prev_yr=m_y,
            year_fact=y_f, year_to_plan=y_p, year_to_prev_yr=y_y,
            populates=populates,
            cell_ref=f"B{r}",
            source_sheet=sheet_name,
            source_row=r,
        ))

        if has_comment:
            # commentary is also stored in report_comments (separated, searchable)
            comments.append(dict(
                indicator_number=num_str,
                commentary=comm,
                management_action=action,
                row_index=r,
                metric_cell_ref=f"B{r}",
            ))

        if depth:
            stack.append((depth, num_str, name))

    return metrics, comments


def parse_investment(ws, sheet_name, is_forecast, header_row=4, data_start=8):
    """Parse Инвест / Инвест Факт sheet → investment_metrics dicts.

    Column layout:
      B=item_no  C=program/title  D=federal_project  E=code_spiui  F=zone
      G..K = инвест-затраты {утв_план_года, план_периода, факт/прогноз, %к плану периода, %к плану года}
      L..P = ввод-фондов     {утв_план_года, план_периода, факт/прогноз, %к плану периода, %к плану года}
      Q=commentary  R=management_action
    """
    rows, comments = [], []
    for r in range(data_start, ws.max_row + 1):
        get = lambda c: ws.cell(row=r, column=c).value
        program = clean(get(3))
        if not program:
            continue
        fed = clean(get(4))
        code = clean(get(5))
        zone = get(6)
        zone_int = int(zone) if isinstance(zone, (int, float)) and int(zone) in (0, 1, 2, 4) else None

        rows.append(dict(
            code_spiui=code,
            federal_project=fed,
            section_roman=None,
            parent_indicator=None,
            program=program,
            is_forecast=is_forecast,
            zone=zone_int,
            exp_approved_year=num(get(7)),
            exp_period_plan=num(get(8)),
            exp_fact_or_forecast=num(get(9)),
            exp_pct_to_period=num(get(10)),
            exp_pct_to_year=num(get(11)),
            funds_approved_year=num(get(12)),
            funds_period_plan=num(get(13)),
            funds_fact_or_forecast=num(get(14)),
            funds_pct_to_period=num(get(15)),
            funds_pct_to_year=num(get(16)),
            cell_ref=f"C{r}",
            source_sheet=sheet_name,
            source_row=r,
        ))
        comm = clean(get(17))
        action = clean(get(18))
        if comm or action:
            comments.append(dict(
                indicator_number=code,
                commentary=comm,
                management_action=action,
                row_index=r,
                metric_cell_ref=f"C{r}",
            ))
    return rows, comments


# Sheet dispatch
SHEET_HANDLERS = [
    ("Доклад Ц ЦЗ",   "I",            parse_operational, dict(header_row=4, data_start=6)),
    ("Срок доставки", "II",           parse_operational, dict(header_row=4, data_start=6)),
    ("Инвест",        "III-forecast", parse_investment,  dict(header_row=6, data_start=8)),
    ("Инвест Факт",   "III-actual",   parse_investment,  dict(header_row=6, data_start=8)),
]


# ---------------------------------------------------------------------------
# DB writes
# ---------------------------------------------------------------------------
def upsert_report(conn, file_path, report_date, sha256, sheets_count, force=False):
    """Insert a report row (or skip if sha256 already loaded and not --force).
    Returns (report_id, was_new). If force=True and any report exists for the
    date, delete it first (cascades to sheets/metrics/comments/investment)."""
    cur = conn.cursor()
    if force:
        cur.execute("DELETE FROM reports WHERE report_date = %s", (report_date,))
    cur.execute("SELECT id FROM reports WHERE sha256 = %s", (sha256,))
    row = cur.fetchone()
    if row:
        return row[0], False
    cur.execute(
        """INSERT INTO reports (filename, report_date, sha256, sheets_count, file_path)
           VALUES (%s, %s, %s, %s, %s)
           RETURNING id""",
        (os.path.basename(file_path), report_date, sha256, sheets_count, file_path),
    )
    return cur.fetchone()[0], True


def write_sheet(conn, report_id, sheet_name, sheet_index, row_count, col_count):
    cur = conn.cursor()
    cur.execute(
        """INSERT INTO report_sheets (report_id, sheet_name, sheet_index, row_count, col_count)
           VALUES (%s, %s, %s, %s, %s) RETURNING id""",
        (report_id, sheet_name, sheet_index, row_count, col_count),
    )
    return cur.fetchone()[0]


def write_metrics(conn, report_id, sheet_id, metrics_list):
    if not metrics_list:
        return []
    cur = conn.cursor()
    out_ids = []
    for m in metrics_list:
        cur.execute(
            """INSERT INTO metrics
               (report_id, sheet_id, indicator_number, parent_indicator, is_priority,
                section_roman, name, category, road, unit, responsible, zone,
                day_fact, day_to_plan, day_to_prev_year,
                month_fact, month_to_plan, month_to_prev_yr,
                year_fact, year_to_plan, year_to_prev_yr,
                populates, cell_ref, source_sheet, source_row)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                       %s,%s,%s,
                       %s,%s,%s,
                       %s,%s,%s,
                       %s,%s,%s,%s)
               RETURNING id""",
            (report_id, sheet_id, m["indicator_number"], m["parent_indicator"], m["is_priority"],
             m["section_roman"], m["name"], m.get("category"), m.get("road"),
             m["unit"], m["responsible"], m["zone"],
             m["day_fact"], m["day_to_plan"], m["day_to_prev_year"],
             m["month_fact"], m["month_to_plan"], m["month_to_prev_yr"],
             m["year_fact"], m["year_to_plan"], m["year_to_prev_yr"],
             m["populates"], m["cell_ref"], m["source_sheet"], m["source_row"]),
        )
        out_ids.append((m["cell_ref"], cur.fetchone()[0]))
    return out_ids  # list of (cell_ref, metric_id) for stitching comments


def write_investment(conn, report_id, sheet_id, rows):
    if not rows:
        return
    cur = conn.cursor()
    for r in rows:
        cur.execute(
            """INSERT INTO investment_metrics
               (report_id, sheet_id, code_spiui, federal_project, section_roman,
                parent_indicator, program, is_forecast, zone,
                exp_approved_year, exp_period_plan, exp_fact_or_forecast,
                exp_pct_to_period, exp_pct_to_year,
                funds_approved_year, funds_period_plan, funds_fact_or_forecast,
                funds_pct_to_period, funds_pct_to_year,
                cell_ref, source_sheet, source_row)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,
                       %s,%s,%s,%s,%s,
                       %s,%s,%s,%s,%s,
                       %s,%s,%s)""",
            (report_id, sheet_id, r["code_spiui"], r["federal_project"], r["section_roman"],
             r["parent_indicator"], r["program"], r["is_forecast"], r["zone"],
             r["exp_approved_year"], r["exp_period_plan"], r["exp_fact_or_forecast"],
             r["exp_pct_to_period"], r["exp_pct_to_year"],
             r["funds_approved_year"], r["funds_period_plan"], r["funds_fact_or_forecast"],
             r["funds_pct_to_period"], r["funds_pct_to_year"],
             r["cell_ref"], r["source_sheet"], r["source_row"]),
        )


def write_comments(conn, report_id, sheet_id, comments_list, metric_id_by_cell):
    if not comments_list:
        return
    cur = conn.cursor()
    for c in comments_list:
        mid = metric_id_by_cell.get(c.get("metric_cell_ref"))
        cur.execute(
            """INSERT INTO report_comments
               (report_id, metric_id, sheet_id, indicator_number, commentary,
                management_action, row_index)
               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (report_id, mid, sheet_id, c.get("indicator_number"),
             c.get("commentary"), c.get("management_action"), c["row_index"]),
        )


def finalize_report(conn, report_id, baseline_year):
    """Compute red/yellow/green counts from inserted rows."""
    cur = conn.cursor()
    cur.execute(
        """SELECT
              count(*) FILTER (WHERE zone = 2),
              count(*) FILTER (WHERE zone = 1),
              count(*) FILTER (WHERE zone = 0),
              count(*)
           FROM metrics WHERE report_id = %s""",
        (report_id,),
    )
    red, yellow, green, total = cur.fetchone()
    cur.execute(
        """UPDATE reports
           SET red_count=%s, yellow_count=%s, green_count=%s,
               metrics_count=%s, baseline_year=%s, updated_at=now()
           WHERE id=%s""",
        (red, yellow, green, total, baseline_year, report_id),
    )
    return dict(red=red, yellow=yellow, green=green, total=total)


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------
def ingest_one(conn, path, force=False, override_date=None):
    name = os.path.basename(path)
    if name.startswith("~$"):
        return None  # Excel lock file
    sha = sha256_file(path)
    report_date = infer_date(path, override_date)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets_present = [s for s, *_ in SHEET_HANDLERS if s in wb.sheetnames]
    baseline_year = detect_baseline_year(wb)

    report_id, is_new = upsert_report(conn, path, report_date, sha, len(sheets_present), force=force)
    if not is_new and not force:
        print(f"  [skip] {name}: same sha256 already loaded (report_id={report_id})")
        wb.close()
        return report_id

    sheets_loaded = []
    for sheet_name, section_code, handler, opts in SHEET_HANDLERS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_id = write_sheet(conn, report_id, sheet_name, list(wb.sheetnames).index(sheet_name),
                               ws.max_row, ws.max_column)

        if handler is parse_operational:
            mlist, clist = parse_operational(ws, sheet_name, section_code, **opts)
            ids = write_metrics(conn, report_id, sheet_id, mlist)
            metric_by_cell = dict(ids)
            write_comments(conn, report_id, sheet_id, clist, metric_by_cell)
            sheets_loaded.append(f"{sheet_name}({len(mlist)} metrics, {len(clist)} comments)")
        else:  # parse_investment
            is_forecast = (sheet_name == "Инвест")
            irows, clist = parse_investment(ws, sheet_name, is_forecast, **opts)
            write_investment(conn, report_id, sheet_id, irows)
            write_comments(conn, report_id, sheet_id, clist, {})
            sheets_loaded.append(f"{sheet_name}({len(irows)} invest, {len(clist)} comments)")

    wb.close()
    stats = finalize_report(conn, report_id, baseline_year)
    conn.commit()
    print(f"  [ok]  {name} → {report_date} (baseline={baseline_year}) "
          f"red={stats['red']} yellow={stats['yellow']} green={stats['green']} "
          f"total={stats['total']}  | {', '.join(sheets_loaded)}")
    return report_id


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src", help="xlsx file or directory")
    ap.add_argument("--date", help="override report_date (YYYY-MM-DD)")
    ap.add_argument("--force", action="store_true",
                    help="delete existing report for the date before inserting")
    args = ap.parse_args()

    paths = []
    if os.path.isdir(args.src):
        for f in sorted(os.listdir(args.src)):
            if f.lower().endswith(".xlsx") and not f.startswith("~$"):
                paths.append(os.path.join(args.src, f))
    else:
        paths = [args.src]

    print(f"Ingesting {len(paths)} file(s)…")
    conn = connect()
    try:
        for p in paths:
            try:
                ingest_one(conn, p, force=args.force, override_date=args.date)
            except Exception as e:
                conn.rollback()
                print(f"  [error] {os.path.basename(p)}: {e}")
    finally:
        conn.close()
    print("done.")


if __name__ == "__main__":
    main()
