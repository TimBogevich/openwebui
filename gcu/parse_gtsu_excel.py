# -*- coding: utf-8 -*-
"""
Parse a ГЦУ (GCU) daily-report Excel and load it into a local PostgreSQL in the
AI-search-friendly `gtsu_search` shape — the same table the cluster has, so the
assistant (Open WebUI + LM Studio + MCP) can query it offline.

This is the standalone, parameterized version of the original cluster ETL. It:
  - parses all 4 ГЦУ sheets (Доклад Ц ЦЗ, Срок доставки, Инвест, Инвест Факт),
  - preserves the indicator hierarchy (item_number / parent_path),
  - keeps every numeric cell in a JSONB `metrics` column (keys in Russian),
  - generates a human-readable `narrative` paragraph per row (Russian FTS + trgm),
  - writes the flat `gtsu_search` table the model queries.

Usage (run with a python that has openpyxl + psycopg/psycopg2):

    set PGHOST=127.0.0.1
    set PGPORT=5432
    set PGUSER=postgres
    set PGPASSWORD=...
    set PGDATABASE=postgres
    python parse_gtsu_excel.py "C:\\path\\to\\ГЦУ-03-31.xlsx"

Optional 2nd arg = report date (YYYY-MM-DD); otherwise inferred from the
filename's MM-DD (year defaults to 2022 to match the «к 2021» comparisons), or
falls back to today.

Idempotent per report_date: re-loading the same date replaces its rows. The
schema/MV are created if missing. After load, `gtsu_search` is (re)built.
"""
import sys
import os
import re
import io
import datetime as dt

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl


# ---------------------------------------------------------------------------
# DB driver shim (psycopg v3 or psycopg2)
# ---------------------------------------------------------------------------
def connect(dsn=None):
    params = dict(
        host=os.environ.get("PGHOST", "127.0.0.1"),
        port=int(os.environ.get("PGPORT", "5432")),
        user=os.environ.get("PGUSER", "postgres"),
        password=os.environ.get("PGPASSWORD", "postgres"),
        dbname=os.environ.get("PGDATABASE", "postgres"),
    )
    try:
        import psycopg
        from psycopg.types.json import Json
        conn = psycopg.connect(dsn, connect_timeout=8) if dsn else psycopg.connect(connect_timeout=8, **params)
        return conn, Json, "psycopg3"
    except ImportError:
        import psycopg2
        from psycopg2.extras import Json
        conn = psycopg2.connect(dsn) if dsn else psycopg2.connect(connect_timeout=8, **params)
        return conn, Json, "psycopg2"


# ---------------------------------------------------------------------------
# Schema — flat gtsu_search table (matches the cluster's AI surface).
# ---------------------------------------------------------------------------
DDL = r"""
CREATE EXTENSION IF NOT EXISTS pg_trgm;
CREATE EXTENSION IF NOT EXISTS unaccent;

CREATE TABLE IF NOT EXISTS gtsu_search (
    id              BIGSERIAL PRIMARY KEY,
    report_date     DATE NOT NULL,
    section_code    TEXT NOT NULL,
    section_title   TEXT NOT NULL,
    sheet_name      TEXT NOT NULL,
    item_number     TEXT,
    item_depth      INT,
    parent_path     TEXT,
    indicator       TEXT NOT NULL,
    full_indicator  TEXT,
    unit            TEXT,
    responsible     TEXT,
    color_marker    INT,
    metrics         JSONB DEFAULT '{}'::jsonb,
    text_comment    TEXT,
    management_actions TEXT,
    narrative       TEXT NOT NULL,
    narrative_tsv   TSVECTOR
        GENERATED ALWAYS AS (to_tsvector('russian', coalesce(narrative,''))) STORED,
    source_row      INT
);

CREATE INDEX IF NOT EXISTS gtsu_search_tsv_idx     ON gtsu_search USING GIN (narrative_tsv);
CREATE INDEX IF NOT EXISTS gtsu_search_trgm_idx    ON gtsu_search USING GIN (narrative gin_trgm_ops);
CREATE INDEX IF NOT EXISTS gtsu_search_date_idx    ON gtsu_search (report_date);
CREATE INDEX IF NOT EXISTS gtsu_search_section_idx ON gtsu_search (section_code);
CREATE INDEX IF NOT EXISTS gtsu_search_item_idx    ON gtsu_search (item_number);
CREATE INDEX IF NOT EXISTS gtsu_search_resp_idx    ON gtsu_search (responsible);

COMMENT ON TABLE gtsu_search IS 'Ежедневный Доклад ГЦУ ОАО РЖД — плоская AI-friendly витрина';
"""


# ---------------------------------------------------------------------------
# Helpers (from the proven cluster ETL)
# ---------------------------------------------------------------------------
def fmt_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if abs(v) < 1 and v != 0:
            return f"{v*100:+.2f}%".replace(".", ",")
        return f"{v:,.3f}".rstrip("0").rstrip(".").replace(",", " ").replace(".", ",")
    return str(v)


def clean(s):
    if s is None:
        return None
    s = str(s).replace("\xa0", " ").strip()
    return s or None


def item_depth(num):
    if not num:
        return None
    return len([p for p in str(num).split(".") if p])


# ---------------------------------------------------------------------------
# Sheet parsers (I/II = Доклад Ц ЦЗ + Срок доставки; III = Инвест + Инвест Факт)
# ---------------------------------------------------------------------------
def parse_section_I_II(ws, section_code, section_title, sheet_name, header_row, data_start_row, report_date):
    rows = list(ws.iter_rows(values_only=True))
    header_cells = rows[header_row - 1]
    col_off = None
    for i, v in enumerate(header_cells):
        if v and "№ показателя" in str(v):
            col_off = i
            break
    if col_off is None:
        col_off = 1

    out, stack, current_block = [], [], None
    for ridx, row in enumerate(rows[data_start_row - 1:], start=data_start_row):
        c = lambda k: row[col_off + k] if col_off + k < len(row) else None
        num, mark, ind = clean(c(0)), clean(c(1)), clean(c(2))
        unit, color, resp = clean(c(3)), c(4), clean(c(5))
        f_d, pl_d, y_d = c(6), c(7), c(8)
        f_m, pl_m, y_m = c(9), c(10), c(11)
        f_y, pl_y, y_y = c(12), c(13), c(14)
        comment, actions = clean(c(15)), clean(c(16))

        if not ind:
            continue
        if re.match(r"^[IVX]+\.\s", ind) and not num:
            current_block, stack = ind, []
            continue
        if num and re.match(r"^\d+$", num) and all(v is None for v in (f_d, pl_d, y_d, f_m, pl_m, y_m, f_y, pl_y, y_y)):
            current_block, stack = ind, [(1, num, ind)]
            continue

        depth = item_depth(num)
        if depth:
            stack = [s for s in stack if s[0] < depth]
        parent_path = " > ".join(s[2] for s in stack) if stack else current_block
        full_indicator = (parent_path + " > " + ind) if parent_path else ind

        metrics = {}
        for k, v in (("факт_сутки", f_d), ("сутки_к_плану", pl_d), ("сутки_к_2021", y_d),
                     ("факт_месяц", f_m), ("месяц_к_плану", pl_m), ("месяц_к_2021", y_m),
                     ("факт_год", f_y), ("год_к_плану", pl_y), ("год_к_2021", y_y)):
            if isinstance(v, (int, float)):
                metrics[k] = v

        bits = [f"Доклад ГЦУ ОАО РЖД от {report_date.strftime('%d.%m.%Y')}.",
                f"Раздел {section_code}. {section_title}."]
        if num:
            bits.append(f"Показатель № {num}.")
        if parent_path:
            bits.append(f"Раздел иерархии: {parent_path}.")
        bits.append(f"Показатель: {ind}.")
        if unit:
            bits.append(f"Единица измерения: {unit}.")
        if resp:
            bits.append(f"Ответственное подразделение: {resp}.")

        def period(label, f, pl, y):
            sub = []
            if isinstance(f, (int, float)):
                sub.append(f"факт {fmt_num(f)}" + (f" {unit}" if unit else ""))
            if isinstance(pl, (int, float)):
                sub.append(f"к плану {fmt_num(pl)}")
            if isinstance(y, (int, float)):
                sub.append(f"к 2021 г. {fmt_num(y)}")
            if sub:
                bits.append(f"{label}: " + ", ".join(sub) + ".")
        period("За сутки", f_d, pl_d, y_d)
        period("С начала месяца", f_m, pl_m, y_m)
        period("С начала года", f_y, pl_y, y_y)
        if comment:
            bits.append(f"Комментарий: {comment}")
        if actions:
            bits.append(f"Предлагаемые управленческие решения: {actions}")

        out.append(dict(
            report_date=report_date, section_code=section_code, section_title=section_title,
            sheet_name=sheet_name, item_number=num, item_depth=depth, parent_path=parent_path,
            indicator=ind, full_indicator=full_indicator, unit=unit, responsible=resp,
            color_marker=int(color) if isinstance(color, (int, float)) else None,
            metrics=metrics, text_comment=comment, management_actions=actions,
            narrative=" ".join(bits), source_row=ridx,
        ))
        if depth:
            stack.append((depth, num, ind))
    return out


def parse_section_III(ws, section_code, section_title, sheet_name, header_row, data_start_row, fact_label, report_date):
    rows = list(ws.iter_rows(values_only=True))
    col_off = 1
    out, last_at_depth = [], {}
    for ridx, row in enumerate(rows[data_start_row - 1:], start=data_start_row):
        c = lambda k: row[col_off + k] if col_off + k < len(row) else None
        title, fed_proj, code, color = clean(c(1)), clean(c(2)), clean(c(3)), c(4)
        iz = [c(5), c(6), c(7), c(8), c(9)]
        vf = [c(10), c(11), c(12), c(13), c(14)]
        comment, actions = clean(c(15)), clean(c(16))
        if not title:
            continue

        depth = item_depth(code) if code else None
        if depth:
            last_at_depth = {d: v for d, v in last_at_depth.items() if d < depth}
        parents = [last_at_depth[d] for d in sorted(last_at_depth)] if last_at_depth else []
        parent_path = " > ".join(parents) if parents else None
        full_indicator = (parent_path + " > " + title) if parent_path else title

        metrics = {}
        keys = ["инвест_затраты_утв_план_года", "инвест_затраты_план_периода",
                f"инвест_затраты_{fact_label}", "инвест_затраты_проц_к_плану_периода",
                "инвест_затраты_проц_к_плану_года", "ввод_фондов_утв_план_года",
                "ввод_фондов_план_периода", f"ввод_фондов_{fact_label}",
                "ввод_фондов_проц_к_плану_периода", "ввод_фондов_проц_к_плану_года"]
        for k, v in zip(keys, iz + vf):
            if isinstance(v, (int, float)):
                metrics[k] = v

        fl = "прогноз" if fact_label == "прогноз" else "факт"
        bits = [f"Доклад ГЦУ ОАО РЖД от {report_date.strftime('%d.%m.%Y')}.",
                f"Раздел {section_code}. {section_title}."]
        if code:
            bits.append(f"Код проекта СПиУИ: {code}.")
        if fed_proj:
            bits.append(f"Федеральный проект: {fed_proj}.")
        if parent_path:
            bits.append(f"Раздел иерархии: {parent_path}.")
        bits.append(f"Программа/Проект: {title}.")
        for label, arr in (("Инвестиционные затраты", iz), ("Ввод фондов", vf)):
            p = []
            names = ["утв. план года", "план периода", fl, "к плану периода", "к плану года"]
            for nm, v in zip(names, arr):
                if isinstance(v, (int, float)):
                    p.append(f"{nm} {fmt_num(v)}" + (" млн ₽" if "план" in nm or nm == fl else ""))
            if p:
                bits.append(f"{label}: " + ", ".join(p) + ".")
        if comment:
            bits.append(f"Комментарий: {comment}")
        if actions:
            bits.append(f"Предлагаемые управленческие решения: {actions}")

        out.append(dict(
            report_date=report_date, section_code=section_code, section_title=section_title,
            sheet_name=sheet_name, item_number=code, item_depth=depth, parent_path=parent_path,
            indicator=title, full_indicator=full_indicator, unit="млн ₽", responsible=fed_proj,
            color_marker=int(color) if isinstance(color, (int, float)) else None,
            metrics=metrics, text_comment=comment, management_actions=actions,
            narrative=" ".join(bits), source_row=ridx,
        ))
        if depth:
            last_at_depth[depth] = title
    return out


# ---------------------------------------------------------------------------
# Report-date inference
# ---------------------------------------------------------------------------
# Open WebUI stores chat uploads as "<uuid>_<original-name>"; strip that prefix
# so the date regex below doesn't latch onto the UUID's digits/dashes.
_UUID_PREFIX = re.compile(
    r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-"
    r"[0-9a-fA-F]{4}-[0-9a-fA-F]{12}_"
)


def infer_date(path, override):
    if override:
        return dt.date.fromisoformat(override)
    name = _UUID_PREFIX.sub("", os.path.basename(path))
    # Full ISO date in the name (e.g. ГЦУ-2026-03-01) wins — respect its year.
    m = re.search(r"(\d{4})[-_.](\d{2})[-_.](\d{2})", name)
    if m:
        y, mm, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return dt.date(y, mm, dd)
        except ValueError:
            pass
    # Short MM-DD form (e.g. ГЦУ-03-09) — default year 2022 («к 2021» base).
    m = re.search(r"(\d{2})[-_.](\d{2})", name)
    if m:
        mm, dd = int(m.group(1)), int(m.group(2))
        try:
            return dt.date(2022, mm, dd)  # «к 2021» comparisons -> 2022 base
        except ValueError:
            pass
    return dt.date.today()


SHEETS = [
    ("Доклад Ц ЦЗ", "I", "Финансово-экономические показатели", "I_II", 1, 3, None),
    ("Срок доставки", "II", "Мониторинг срока доставки грузов и порожних вагонов", "I_II", 4, 6, None),
    ("Инвест", "III-forecast", "Инвестиционная программа ОАО «РЖД» (Прогноз)", "III", 4, 8, "прогноз"),
    ("Инвест Факт", "III-actual", "Инвестиционная программа ОАО «РЖД» (Факт)", "III", 4, 8, "факт"),
]


def main():
    if len(sys.argv) < 2:
        print("usage: python parse_gtsu_excel.py <file.xlsx> [YYYY-MM-DD]")
        sys.exit(1)
    xlsx = sys.argv[1]
    report_date = infer_date(xlsx, sys.argv[2] if len(sys.argv) > 2 else None)
    print(f"Parsing {os.path.basename(xlsx)} as report_date={report_date}")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    rows = []
    for name, code, title, kind, hr, ds, fl in SHEETS:
        if name not in wb.sheetnames:
            print(f"  (sheet '{name}' missing, skipped)")
            continue
        if kind == "I_II":
            rows += parse_section_I_II(wb[name], code, title, name, hr, ds, report_date)
        else:
            rows += parse_section_III(wb[name], code, title, name, hr, ds, fl, report_date)
    print(f"Parsed {len(rows)} rows")

    conn, Json, drv = connect()
    conn.autocommit = False
    cur = conn.cursor()
    print(f"Connected via {drv}. Ensuring schema...")
    cur.execute(DDL)
    # idempotent per date
    cur.execute("DELETE FROM gtsu_search WHERE report_date = %s", (report_date,))
    sql = """INSERT INTO gtsu_search
      (report_date, section_code, section_title, sheet_name, item_number, item_depth,
       parent_path, indicator, full_indicator, unit, responsible, color_marker,
       metrics, text_comment, management_actions, narrative, source_row)
      VALUES (%(report_date)s,%(section_code)s,%(section_title)s,%(sheet_name)s,
              %(item_number)s,%(item_depth)s,%(parent_path)s,%(indicator)s,
              %(full_indicator)s,%(unit)s,%(responsible)s,%(color_marker)s,
              %(metrics)s,%(text_comment)s,%(management_actions)s,%(narrative)s,
              %(source_row)s)"""
    for r in rows:
        r["metrics"] = Json(r["metrics"])
        cur.execute(sql, r)
    conn.commit()
    cur.execute("SELECT count(*), count(DISTINCT section_code) FROM gtsu_search WHERE report_date=%s", (report_date,))
    print("Loaded for this date:", cur.fetchone())
    cur.execute("SELECT section_code, count(*) FROM gtsu_search GROUP BY 1 ORDER BY 1")
    print("Per section (all dates):", cur.fetchall())
    conn.close()
    print("DONE.")


if __name__ == "__main__":
    main()
