# -*- coding: utf-8 -*-
"""
GCU report uploader — standalone web page that ingests a ГЦУ .xlsx straight into
PostgreSQL (table gtsu_search), completely separate from the chat so a 140k-char
report never pollutes the model context.

Flow: drag/drop .xlsx -> validate format -> check if that report_date is already
loaded -> parse into PG -> return one of three statuses:
  • red    — неверный формат файла
  • yellow — файл уже был загружен (дата уже в базе)
  • green  — файл загружен (N строк за <дата>)

Reuses parse_gtsu_excel.py (same logic as gcu-watch). Serves on :8810.
"""
import os
import io
import sys
import tempfile

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from flask import Flask, request, jsonify, Response

import parse_gtsu_excel as P

app = Flask(__name__)

BRAND = "РЖД Интер"


def _date_already_loaded(report_date):
    """Return current row-count for report_date (0 if not loaded)."""
    conn, _, _ = P.connect()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT count(*) FROM gtsu_search WHERE report_date = %s",
                (report_date,),
            )
            return cur.fetchone()[0]
    finally:
        conn.close()


def _ingest(path, original_name):
    """Parse the xlsx at `path` into PG. Returns (status, message, detail)."""
    # 1) format gate — must be .xlsx and openpyxl-loadable with ≥1 GCU sheet
    if not original_name.lower().endswith(".xlsx"):
        return ("error", "Невозможно загрузить: неверный формат файла (нужен .xlsx).", "")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheetnames = set(wb.sheetnames)
        wb.close()
    except Exception as e:
        return ("error", "Невозможно загрузить: файл повреждён или не является Excel.", str(e)[:200])

    known = {name for (name, *_rest) in P.SHEETS}
    if not (sheetnames & known):
        return ("error",
                "Невозможно загрузить: это не доклад ГЦУ (нет ожидаемых листов).",
                "Листы: " + ", ".join(list(sheetnames)[:6]))

    # 2) infer report date the same way the parser/watcher does
    report_date = P.infer_date(original_name, None)

    # 3) duplicate check
    existing = _date_already_loaded(report_date)
    if existing > 0:
        return ("warn",
                f"Файл уже был загружен: доклад за {report_date.strftime('%d.%m.%Y')} "
                f"уже в базе ({existing} строк).",
                "Повторная загрузка перезапишет эти строки — используйте «Загрузить повторно».")

    # 4) parse + load (parse_gtsu_excel.main reads argv; call its pieces directly)
    rows = _parse_rows(path, report_date)
    if not rows:
        return ("error", "Невозможно загрузить: не удалось разобрать данные доклада.", "")
    n = _write_rows(rows, report_date)
    return ("ok",
            f"Файл загружен в базу: {n} строк за {report_date.strftime('%d.%m.%Y')}.",
            f"Источник: {original_name}")


def _parse_rows(xlsx, report_date):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    rows = []
    for name, code, title, kind, hr, ds, fl in P.SHEETS:
        if name not in wb.sheetnames:
            continue
        if kind == "I_II":
            rows += P.parse_section_I_II(wb[name], code, title, name, hr, ds, report_date)
        else:
            rows += P.parse_section_III(wb[name], code, title, name, hr, ds, fl, report_date)
    return rows


def _write_rows(rows, report_date):
    conn, Json, drv = P.connect()
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute(P.DDL)
    cur.execute("DELETE FROM gtsu_search WHERE report_date = %s", (report_date,))
    sql = """INSERT INTO gtsu_search
      (report_date, section_code, section_title, sheet_name, item_number, item_depth,
       parent_path, indicator, full_indicator, unit, responsible, color_marker,
       metrics, text_comment, management_actions, narrative, source_row)
      VALUES (%(report_date)s,%(section_code)s,%(section_title)s,%(sheet_name)s,
              %(item_number)s,%(item_depth)s,%(parent_path)s,%(indicator)s,
              %(full_indicator)s,%(unit)s,%(responsible)s,%(color_marker)s,
              %(metrics)s,%(text_comment)s,%(management_actions)s,%(narrative)s,
              %(source_row)s)"""
    for r in rows:
        r["metrics"] = Json(r["metrics"])
        cur.execute(sql, r)
    conn.commit()
    cur.execute("SELECT count(*) FROM gtsu_search WHERE report_date=%s", (report_date,))
    n = cur.fetchone()[0]
    conn.close()
    return n


@app.route("/api/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify(status="error", message="Файл не выбран."), 400
    force = request.form.get("force") == "1"
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name
    try:
        if force:
            # bypass the duplicate gate: parse + overwrite
            rd = P.infer_date(f.filename, None)
            rows = _parse_rows(tmp_path, rd)
            if not rows:
                return jsonify(status="error", message="Не удалось разобрать данные доклада.")
            n = _write_rows(rows, rd)
            return jsonify(status="ok",
                           message=f"Файл перезагружен: {n} строк за {rd.strftime('%d.%m.%Y')}.",
                           detail=f"Источник: {f.filename}")
        status, message, detail = _ingest(tmp_path, f.filename)
        return jsonify(status=status, message=message, detail=detail)
    except Exception as e:
        return jsonify(status="error", message="Ошибка при загрузке.", detail=str(e)[:300])
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


@app.route("/api/loaded", methods=["GET"])
def loaded():
    """List which report dates are already in the DB (for the page summary)."""
    conn, _, _ = P.connect()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT count(DISTINCT report_date), min(report_date), max(report_date), count(*) "
                        "FROM gtsu_search")
            dates, dmin, dmax, total = cur.fetchone()
        return jsonify(dates=dates, min=str(dmin) if dmin else None,
                       max=str(dmax) if dmax else None, total=total)
    finally:
        conn.close()


@app.route("/static/<path:fname>", methods=["GET"])
def static_file(fname):
    safe = os.path.basename(fname)
    path = os.path.join(os.path.dirname(__file__), "static", safe)
    if not os.path.isfile(path):
        return Response(status=404)
    mt = "image/png" if safe.lower().endswith(".png") else "application/octet-stream"
    with open(path, "rb") as fh:
        return Response(fh.read(), mimetype=mt)


@app.route("/", methods=["GET"])
def index():
    return Response(PAGE, mimetype="text/html; charset=utf-8")


PAGE = """<!doctype html>
<html lang="ru"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Загрузка докладов ГЦУ — РЖД Интер</title>
<style>
  :root{ --rzd-red:#E21A1A; --bg:#3a3a3e; --bg2:#2f2f33; --line:rgba(255,255,255,.10);
         --text:#f2f2f4; --muted:#a6a6ad; --ok:#2bb24c; --warn:#f0a020; --err:#e23b3b; }
  *{box-sizing:border-box}
  body{margin:0;font-family:'Segoe UI',Roboto,Arial,sans-serif;background:var(--bg);
       color:var(--text);min-height:100vh;display:flex;align-items:flex-start;justify-content:center}
  .wrap{width:100%;max-width:620px;padding:40px 24px}
  .head{display:flex;align-items:center;gap:14px;margin-bottom:6px}
  .head img{height:42px}
  .head .word{font-weight:800;font-size:22px;letter-spacing:.12em;
       border-left:1px solid var(--line);padding-left:13px}
  h1{font-size:19px;font-weight:600;margin:22px 0 4px}
  .sub{color:var(--muted);font-size:13px;margin-bottom:22px}
  .drop{border:2px dashed var(--line);border-radius:14px;background:var(--bg2);
        padding:46px 20px;text-align:center;cursor:pointer;transition:.15s}
  .drop.over{border-color:var(--rzd-red);background:#34343a}
  .drop .big{font-size:15px;margin-bottom:6px}
  .drop .small{color:var(--muted);font-size:12px}
  input[type=file]{display:none}
  .msg{margin-top:20px;padding:14px 16px;border-radius:11px;font-size:14px;display:none;
       border:1px solid var(--line);line-height:1.45}
  .msg .detail{display:block;color:var(--muted);font-size:12px;margin-top:5px}
  .msg.ok{background:rgba(43,178,76,.12);border-color:rgba(43,178,76,.4)}
  .msg.ok b{color:var(--ok)}
  .msg.warn{background:rgba(240,160,32,.12);border-color:rgba(240,160,32,.4)}
  .msg.warn b{color:var(--warn)}
  .msg.err{background:rgba(226,59,59,.12);border-color:rgba(226,59,59,.4)}
  .msg.err b{color:var(--err)}
  .reload{margin-top:10px;display:none}
  .reload button{background:var(--warn);color:#1a1a1a;border:0;border-radius:8px;
       padding:8px 14px;font-weight:600;cursor:pointer;font-size:13px}
  .stat{margin-top:26px;color:var(--muted);font-size:12px;text-align:center}
  .spin{display:inline-block;width:14px;height:14px;border:2px solid var(--line);
       border-top-color:var(--rzd-red);border-radius:50%;animation:s .7s linear infinite;
       vertical-align:-2px;margin-right:7px}
  @keyframes s{to{transform:rotate(360deg)}}
</style></head><body>
<div class="wrap">
  <div class="head">
    <img src="/static/rzd_logo_dark.png" alt="РЖД" onerror="this.style.display='none'">
    <span class="word">ИНТЕР</span>
  </div>
  <h1>Загрузка докладов ГЦУ</h1>
  <div class="sub">Перетащите файл доклада (.xlsx) — он будет разобран и загружен в базу.
      Файл не попадает в чат и не влияет на модель.</div>

  <label class="drop" id="drop">
    <div class="big">Перетащите .xlsx сюда или нажмите для выбора</div>
    <div class="small">Принимаются файлы вида ГЦУ-MM-DD.xlsx</div>
    <input type="file" id="file" accept=".xlsx">
  </label>

  <div class="msg" id="msg"><b id="msgtext"></b><span class="detail" id="msgdetail"></span></div>
  <div class="reload" id="reloadBox">
    <button id="reloadBtn">Загрузить повторно (перезаписать)</button>
  </div>

  <div class="stat" id="stat"></div>
</div>
<script>
  var drop=document.getElementById('drop'), fileIn=document.getElementById('file'),
      msg=document.getElementById('msg'), msgt=document.getElementById('msgtext'),
      msgd=document.getElementById('msgdetail'), stat=document.getElementById('stat'),
      reloadBox=document.getElementById('reloadBox'), reloadBtn=document.getElementById('reloadBtn');
  var lastFile=null;

  function show(status, message, detail){
    msg.className='msg '+(status==='ok'?'ok':status==='warn'?'warn':'err');
    msg.style.display='block'; msgt.textContent=message; msgd.textContent=detail||'';
    reloadBox.style.display = (status==='warn') ? 'block' : 'none';
  }
  function send(file, force){
    lastFile=file;
    msg.className='msg'; msg.style.display='block'; reloadBox.style.display='none';
    msgt.innerHTML='<span class="spin"></span>Загрузка…'; msgd.textContent='';
    var fd=new FormData(); fd.append('file',file); if(force) fd.append('force','1');
    fetch('/api/upload',{method:'POST',body:fd}).then(function(r){return r.json()}).then(function(d){
      show(d.status, d.message, d.detail); refreshStat();
    }).catch(function(e){ show('error','Ошибка сети при загрузке.', String(e)); });
  }
  function refreshStat(){
    fetch('/api/loaded').then(function(r){return r.json()}).then(function(d){
      if(d.dates) stat.textContent='В базе: '+d.total+' строк, '+d.dates+' дат ('+d.min+' … '+d.max+').';
    }).catch(function(){});
  }
  drop.addEventListener('click',function(){fileIn.click()});
  fileIn.addEventListener('change',function(){ if(fileIn.files[0]) send(fileIn.files[0],false); });
  ['dragover','dragenter'].forEach(function(ev){drop.addEventListener(ev,function(e){
    e.preventDefault(); drop.classList.add('over');});});
  ['dragleave','drop'].forEach(function(ev){drop.addEventListener(ev,function(e){
    e.preventDefault(); drop.classList.remove('over');});});
  drop.addEventListener('drop',function(e){ var f=e.dataTransfer.files[0]; if(f) send(f,false); });
  reloadBtn.addEventListener('click',function(){ if(lastFile) send(lastFile,true); });
  refreshStat();
</script>
</body></html>"""


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8810"))
    print(f"GCU uploader on http://0.0.0.0:{port}")
    app.run(host="0.0.0.0", port=port)
